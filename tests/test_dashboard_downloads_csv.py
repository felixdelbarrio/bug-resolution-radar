from __future__ import annotations

from io import BytesIO

import pandas as pd
from openpyxl import load_workbook

from bug_resolution_radar.ui.dashboard.downloads import df_to_csv_bytes, df_to_excel_bytes


def test_df_to_csv_bytes_encodes_key_as_hyperlink_formula_when_url_exists() -> None:
    df = pd.DataFrame(
        [
            {
                "key": "INC-123",
                "summary": "Error login",
                "url": "https://jira.example.com/browse/INC-123",
            }
        ]
    )

    csv_txt = df_to_csv_bytes(df).decode("utf-8")

    assert '"=HYPERLINK(""https://jira.example.com/browse/INC-123"",""INC-123"")"' in csv_txt
    assert "url\n" not in csv_txt


def test_df_to_csv_bytes_keeps_plain_key_when_url_is_not_http() -> None:
    df = pd.DataFrame([{"key": "INC-1", "url": "not-a-url"}])

    csv_txt = df_to_csv_bytes(df).decode("utf-8")

    assert "INC-1" in csv_txt
    assert "HYPERLINK(" not in csv_txt
    assert "url\n" not in csv_txt


def test_df_to_excel_bytes_makes_key_clickable_and_hides_url_column() -> None:
    df = pd.DataFrame(
        [
            {
                "key": "INC000104226433",
                "summary": "Incidencia",
                "url": "https://itsmhelixbbva-smartit.onbmc.com/smartit/app/#/incidentPV/ABC123",
            }
        ]
    )

    xlsx = df_to_excel_bytes(df, sheet_name="Issues")

    wb = load_workbook(BytesIO(xlsx))
    ws = wb["Issues"]

    headers = [c.value for c in ws[1]]
    assert "url" not in headers
    assert headers[:2] == ["key", "summary"]

    cell = ws["A2"]
    assert cell.value == "INC000104226433"
    assert cell.hyperlink is not None
    assert (
        str(cell.hyperlink.target)
        == "https://itsmhelixbbva-smartit.onbmc.com/smartit/app/#/incidentPV/ABC123"
    )


def test_df_to_excel_bytes_accepts_timezone_aware_datetimes() -> None:
    df = pd.DataFrame(
        [
            {
                "key": "INC-1",
                "updated": pd.Timestamp("2026-02-25T10:03:03Z"),
            }
        ]
    )

    xlsx = df_to_excel_bytes(df, sheet_name="Issues")

    wb = load_workbook(BytesIO(xlsx))
    ws = wb["Issues"]
    assert ws["B2"].value is not None
    assert ws["B2"].number_format.lower() == "dd/mm/yyyy hh:mm:ss"
