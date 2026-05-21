from __future__ import annotations

from io import BytesIO
from typing import Any

import pandas as pd
from openpyxl import load_workbook

from bug_resolution_radar.analytics.filtering import FilterState
from bug_resolution_radar.config import Settings
from bug_resolution_radar.services import issue_workbook_export
from bug_resolution_radar.services.dashboard_snapshot import (
    DashboardQuery,
    DashboardScopeContext,
)
from bug_resolution_radar.services.workspace import WorkspaceSelection


def test_finalist_discrepancies_workbook_export(monkeypatch: Any, tmp_path) -> None:
    query = DashboardQuery(
        workspace=WorkspaceSelection(country="México", source_id="jira:mexico:senda"),
        filters=FilterState(status=[], priority=[], assignee=[]),
    )
    discrepancies = pd.DataFrame(
        [
            {
                "helix_id": "INC000104154954",
                "helix_status": "Closed",
                "helix_url": "https://helix.example.com/INC000104154954",
                "jira_key": "MEX-1",
                "jira_summary": "Pendiente en JIRA",
                "jira_status": "To Rework",
                "jira_priority": "High",
                "jira_assignee": "Ana",
                "jira_open_days": 20,
                "jira_url": "https://jira.example.com/browse/MEX-1",
                "source_alias": "Senda",
            }
        ]
    )
    context = DashboardScopeContext(
        scoped_df=pd.DataFrame(),
        dff=pd.DataFrame(),
        open_df=pd.DataFrame(),
        source_ids=("jira:mexico:senda",),
        kpis={},
        finalist_discrepancies=discrepancies,
    )
    monkeypatch.setattr(
        issue_workbook_export,
        "load_scope_context",
        lambda settings, *, query: context,
    )

    export = issue_workbook_export.build_finalist_discrepancies_workbook_export(
        Settings(DATA_PATH=str(tmp_path / "issues.json")),
        query=query,
    )

    assert export.row_count == 1
    wb = load_workbook(BytesIO(export.content))
    ws = wb["Discrepancias finalistas"]
    headers = [cell.value for cell in ws[1]]
    assert headers[:5] == ["Helix ID", "Estado Helix", "JIRA key", "Resumen JIRA", "Estado JIRA"]
    assert ws["A2"].value == "INC000104154954"
    assert ws["C2"].value == "MEX-1"
