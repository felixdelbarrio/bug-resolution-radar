"""Reusable export controls and helper serializers for dashboard downloads."""

from __future__ import annotations

import html
from dataclasses import dataclass
from datetime import date, datetime, timezone
from io import BytesIO
from typing import Any, Iterable, Literal, Optional, Sequence

import pandas as pd
import streamlit as st
from openpyxl.utils import get_column_letter

from bug_resolution_radar.theme.design_tokens import BBVA_FONT_HEADLINE, BBVA_FONT_SANS, BBVA_LIGHT

EXCEL_DATETIME_NUMFMT = "dd/mm/yyyy hh:mm:ss"
EXCEL_DEFAULT_DATA_ROW_HEIGHT = 18.0
EXCEL_DEFAULT_HEADER_ROW_HEIGHT = 20.0
EXCEL_ID_COL_MIN_WIDTH = 18.0
EXCEL_ID_COL_MAX_WIDTH = 26.0


# -------------------------
# CSV helpers
# -------------------------
@dataclass(frozen=True)
class CsvDownloadSpec:
    filename_prefix: str = "issues_filtradas"
    include_index: bool = False
    encoding: str = "utf-8"
    mime: str = ""
    date_format: str = "%Y-%m-%d"
    format: Literal["xlsx", "csv"] = "xlsx"


def _safe_filename(s: str) -> str:
    s = (s or "").strip().replace(" ", "_")
    keep = []
    for ch in s:
        if ch.isalnum() or ch in {"_", "-", "."}:
            keep.append(ch)
    return "".join(keep) or "export"


def _timestamp() -> str:
    # local-ish timestamp string; good enough for filenames
    return datetime.now().strftime("%Y%m%d_%H%M%S")


def _build_filename(prefix: str, *, suffix: str = "", ext: str = "csv") -> str:
    pref = _safe_filename(prefix)
    suf = _safe_filename(suffix) if suffix else ""
    ts = _timestamp()
    dot_ext = (ext or "").strip().lstrip(".") or "txt"
    return f"{pref}{'_' + suf if suf else ''}_{ts}.{dot_ext}"


def build_download_filename(prefix: str, *, suffix: str = "", ext: str = "csv") -> str:
    """Public wrapper for timestamped export filenames."""
    return _build_filename(prefix, suffix=suffix, ext=ext)


def _csv_hyperlink_formula(url: object, label: object) -> str | None:
    url_txt = str(url or "").strip()
    if not url_txt:
        return None
    if not (url_txt.startswith("http://") or url_txt.startswith("https://")):
        return None

    label_txt = str(label or "").strip() or url_txt
    safe_url = url_txt.replace('"', '""')
    safe_label = label_txt.replace('"', '""')
    return f'=HYPERLINK("{safe_url}","{safe_label}")'


def _csv_key_as_link_df(df: pd.DataFrame) -> pd.DataFrame:
    """Return a CSV-safe copy where `key` becomes a hyperlink formula when `url` exists."""
    if df is None or df.empty:
        return pd.DataFrame() if df is None else df.copy()
    if "key" not in df.columns or "url" not in df.columns:
        return df

    out = df.copy(deep=False)
    key_values = out["key"].tolist()
    url_values = out["url"].tolist()
    link_values = [
        _csv_hyperlink_formula(url=url, label=key) for key, url in zip(key_values, url_values)
    ]
    out["key"] = [
        link if link is not None else str(orig if orig is not None else "")
        for link, orig in zip(link_values, key_values)
    ]
    # Once the link is embedded in `key`, the raw `url` column is redundant in the CSV export.
    out = out.drop(columns=["url"], errors="ignore")
    return out


def _xlsx_link_specs_from_df(
    df: pd.DataFrame,
    *,
    default_visible_col: str = "key",
    default_url_col: str = "url",
) -> list[tuple[str, str]]:
    if df is None or df.empty:
        return []
    if default_visible_col in df.columns and default_url_col in df.columns:
        return [(default_visible_col, default_url_col)]
    return []


def _prepare_excel_df_for_links(
    df: pd.DataFrame,
    *,
    link_specs: Sequence[tuple[str, str]],
) -> pd.DataFrame:
    if df is None:
        return pd.DataFrame()
    out = df.copy()
    url_cols_to_drop: list[str] = []
    for visible_col, url_col in link_specs:
        if visible_col in out.columns and url_col in out.columns:
            url_cols_to_drop.append(url_col)
    if url_cols_to_drop:
        out = out.drop(columns=url_cols_to_drop, errors="ignore")
    return out


def _excel_safe_scalar(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, pd.Timestamp):
        if pd.isna(value):
            return None
        if value.tzinfo is not None:
            return value.tz_convert("UTC").tz_localize(None).to_pydatetime()
        return value.to_pydatetime()
    if isinstance(value, datetime):
        if value.tzinfo is not None:
            return value.astimezone(timezone.utc).replace(tzinfo=None)
        return value
    return value


def _excel_safe_df(df: pd.DataFrame) -> pd.DataFrame:
    if df is None or df.empty:
        return pd.DataFrame() if df is None else df.copy()

    out = df.copy()
    for col in out.columns:
        series = out[col]
        try:
            if isinstance(series.dtype, pd.DatetimeTZDtype):
                out[col] = series.dt.tz_convert("UTC").dt.tz_localize(None)
                continue
            if pd.api.types.is_object_dtype(series.dtype):
                out[col] = series.map(_excel_safe_scalar)
        except Exception:
            # Export path must be resilient even if a column has mixed/unexpected values.
            out[col] = (
                series.map(_excel_safe_scalar)
                if pd.api.types.is_object_dtype(series.dtype)
                else series
            )
    return out


def _safe_excel_sheet_name(name: str, *, used: set[str]) -> str:
    base = str(name or "Export").strip() or "Export"
    cleaned = "".join(ch for ch in base if ch not in {":", "\\", "/", "?", "*", "[", "]"})
    cleaned = cleaned[:31] or "Export"
    candidate = cleaned
    n = 2
    while candidate in used:
        suffix = f"_{n}"
        candidate = (cleaned[: max(1, 31 - len(suffix))] + suffix)[:31]
        n += 1
    used.add(candidate)
    return candidate


def _normalize_link_specs_cache_key(
    hyperlink_columns_by_sheet: Optional[dict[str, Sequence[tuple[str, str]]]],
) -> tuple[tuple[str, tuple[tuple[str, str], ...]], ...]:
    if not hyperlink_columns_by_sheet:
        return ()
    out: list[tuple[str, tuple[tuple[str, str], ...]]] = []
    for sheet_name in sorted(hyperlink_columns_by_sheet.keys()):
        specs = hyperlink_columns_by_sheet.get(sheet_name) or ()
        out.append(
            (
                str(sheet_name),
                tuple((str(visible), str(url_col)) for visible, url_col in specs),
            )
        )
    return tuple(out)


def _excel_text_len(value: Any) -> int:
    if value is None:
        return 0
    return len(str(value))


def _find_excel_id_column(df: pd.DataFrame) -> Optional[str]:
    if df is None:
        return None
    for name in ("ID de la Incidencia", "key", "id"):
        if name in df.columns:
            return name
    return None


def _apply_excel_id_sizing(ws: Any, out_df: pd.DataFrame) -> None:
    """Size workbook layout using the incident ID column instead of long text columns."""
    if out_df is None or out_df.empty:
        return

    id_col_name = _find_excel_id_column(out_df)
    if not id_col_name:
        return

    try:
        id_col_idx = int(out_df.columns.get_loc(id_col_name)) + 1
    except Exception:
        return

    col_letter = get_column_letter(id_col_idx)
    max_chars = _excel_text_len(id_col_name)
    for row_idx in range(2, ws.max_row + 1):
        cell_val = ws.cell(row=row_idx, column=id_col_idx).value
        max_chars = max(max_chars, _excel_text_len(cell_val))

    target_width = min(EXCEL_ID_COL_MAX_WIDTH, max(EXCEL_ID_COL_MIN_WIDTH, float(max_chars + 2)))
    ws.column_dimensions[col_letter].width = float(target_width)

    ws.row_dimensions[1].height = EXCEL_DEFAULT_HEADER_ROW_HEIGHT
    for row_idx in range(2, ws.max_row + 1):
        id_val = ws.cell(row=row_idx, column=id_col_idx).value
        line_count = max(1, str(id_val or "").count("\n") + 1)
        ws.row_dimensions[row_idx].height = EXCEL_DEFAULT_DATA_ROW_HEIGHT * float(line_count)


def _write_excel_sheet(
    writer: pd.ExcelWriter,
    *,
    sheet_name: str,
    df: pd.DataFrame,
    include_index: bool,
    hyperlink_columns: Optional[Sequence[tuple[str, str]]] = None,
) -> None:
    src_df = pd.DataFrame() if df is None else df
    link_specs = (
        list(hyperlink_columns)
        if hyperlink_columns is not None
        else _xlsx_link_specs_from_df(src_df)
    )
    out_df = _prepare_excel_df_for_links(src_df, link_specs=link_specs)
    out_df = _excel_safe_df(out_df)

    out_df.to_excel(writer, index=include_index, sheet_name=sheet_name)
    ws = writer.book[sheet_name]
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            if isinstance(cell.value, datetime):
                cell.number_format = EXCEL_DATETIME_NUMFMT
            elif isinstance(cell.value, date):
                cell.number_format = EXCEL_DATETIME_NUMFMT
    _apply_excel_id_sizing(ws, out_df)
    if out_df.empty or not link_specs:
        return

    col_positions = {str(col): idx + 1 for idx, col in enumerate(out_df.columns.tolist())}
    row_offset = 2
    for visible_col, url_col in link_specs:
        if visible_col not in src_df.columns or url_col not in src_df.columns:
            continue
        target_col = col_positions.get(str(visible_col))
        if target_col is None:
            continue
        labels = src_df[visible_col].tolist()
        urls = src_df[url_col].tolist()
        for row_idx, (label, url) in enumerate(zip(labels, urls), start=row_offset):
            url_txt = str(url or "").strip()
            if not (url_txt.startswith("http://") or url_txt.startswith("https://")):
                continue
            cell = ws.cell(row=row_idx, column=target_col)
            cell.value = str(label or "").strip() or url_txt
            cell.hyperlink = url_txt
            cell.style = "Hyperlink"


@st.cache_data(show_spinner=False, max_entries=64)
def _dfs_to_excel_bytes_cached(
    sheets: Sequence[tuple[str, pd.DataFrame]],
    *,
    include_index: bool = False,
    hyperlink_columns_cache_key: tuple[tuple[str, tuple[tuple[str, str], ...]], ...] = (),
) -> bytes:
    hyperlinks: dict[str, Sequence[tuple[str, str]]] | None = None
    if hyperlink_columns_cache_key:
        hyperlinks = {sheet_name: list(specs) for sheet_name, specs in hyperlink_columns_cache_key}

    bio = BytesIO()
    with pd.ExcelWriter(bio, engine="openpyxl") as writer:
        used_sheet_names: set[str] = set()
        if not sheets:
            _write_excel_sheet(
                writer,
                sheet_name=_safe_excel_sheet_name("Export", used=used_sheet_names),
                df=pd.DataFrame(),
                include_index=include_index,
                hyperlink_columns=None,
            )
        for raw_sheet_name, df in sheets:
            safe_sheet = _safe_excel_sheet_name(raw_sheet_name, used=used_sheet_names)
            sheet_links = None
            if hyperlinks:
                sheet_links = hyperlinks.get(raw_sheet_name)
                if sheet_links is None:
                    sheet_links = hyperlinks.get(safe_sheet)
            _write_excel_sheet(
                writer,
                sheet_name=safe_sheet,
                df=df,
                include_index=include_index,
                hyperlink_columns=sheet_links,
            )
    return bio.getvalue()


def dfs_to_excel_bytes(
    sheets: Sequence[tuple[str, pd.DataFrame]],
    *,
    include_index: bool = False,
    hyperlink_columns_by_sheet: Optional[dict[str, Sequence[tuple[str, str]]]] = None,
) -> bytes:
    """Convert multiple DataFrames into a single XLSX workbook."""
    safe_sheets = tuple((str(name), pd.DataFrame() if df is None else df) for name, df in sheets)
    return _dfs_to_excel_bytes_cached(
        safe_sheets,
        include_index=include_index,
        hyperlink_columns_cache_key=_normalize_link_specs_cache_key(hyperlink_columns_by_sheet),
    )


def df_to_excel_bytes(
    df: pd.DataFrame,
    *,
    include_index: bool = False,
    sheet_name: str = "Issues",
    hyperlink_columns: Optional[Sequence[tuple[str, str]]] = None,
) -> bytes:
    """Convert DataFrame to XLSX bytes.

    Hyperlinks are rendered as real spreadsheet links (not formulas shown as text).
    By default, when columns `key` and `url` exist, `key` is clickable and `url` is hidden.
    """
    src_df = pd.DataFrame() if df is None else df
    return dfs_to_excel_bytes(
        [(sheet_name, src_df)],
        include_index=include_index,
        hyperlink_columns_by_sheet=(
            {str(sheet_name): list(hyperlink_columns)} if hyperlink_columns is not None else None
        ),
    )


def df_to_csv_bytes(
    df: pd.DataFrame, *, include_index: bool = False, encoding: str = "utf-8"
) -> bytes:
    """Convert DataFrame to CSV bytes with safe defaults.

    - Dates are serialized by pandas; keep it simple and consistent.
    """
    if df is None:
        df = pd.DataFrame()
    else:
        df = _csv_key_as_link_df(df)
    csv: str = df.to_csv(index=include_index)
    return csv.encode(encoding, errors="replace")


def download_button_for_df(
    df: pd.DataFrame,
    *,
    label: str = "Descargar Excel",
    key: str = "download_csv",
    spec: Optional[CsvDownloadSpec] = None,
    suffix: str = "",
    disabled: Optional[bool] = None,
    width: Literal["stretch", "content"] | int = "stretch",
) -> None:
    """Render a download button for a dataframe.

    `disabled` defaults to True if df is empty.
    """
    if df is None:
        df = pd.DataFrame()

    is_empty = df.empty
    if disabled is None:
        disabled = is_empty

    csv_spec = spec or CsvDownloadSpec()

    fmt = str(getattr(csv_spec, "format", "xlsx") or "xlsx").strip().lower()
    if fmt not in {"xlsx", "csv"}:
        fmt = "xlsx"

    fname = _build_filename(csv_spec.filename_prefix, suffix=suffix, ext=fmt)
    if fmt == "csv":
        payload = df_to_csv_bytes(
            df, include_index=csv_spec.include_index, encoding=csv_spec.encoding
        )
        mime = csv_spec.mime or "text/csv"
    else:
        payload = df_to_excel_bytes(df, include_index=csv_spec.include_index, sheet_name="Issues")
        mime = csv_spec.mime or "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    st.download_button(
        label=label,
        data=payload,
        file_name=fname,
        mime=mime,
        key=key,
        disabled=disabled,
        width=width,
    )


# -------------------------
# "Table view" shaping
# -------------------------
DEFAULT_TABLE_COLS = [
    "key",
    "summary",
    "status",
    "type",
    "priority",
    "assignee",
    "created",
    "updated",
    "resolved",
    "resolution",
    "url",
]


def make_table_export_df(
    df: pd.DataFrame, *, preferred_cols: Optional[Iterable[str]] = None
) -> pd.DataFrame:
    """Return a dataframe shaped like the table view export.

    - Keeps only known/important columns first
    - Preserves extra columns at the end
    - Does not mutate the input df
    """
    if df is None or df.empty:
        return pd.DataFrame()

    dff = df.copy()

    cols_pref = list(preferred_cols) if preferred_cols is not None else DEFAULT_TABLE_COLS
    existing_pref = [c for c in cols_pref if c in dff.columns]
    extra = [c for c in dff.columns if c not in existing_pref]

    out = dff[existing_pref + extra].copy()

    # Keep datetimes readable; pandas will serialize fine, but normalize tz a bit if present
    for c in ["created", "updated", "resolved"]:
        if c in out.columns:
            try:
                out[c] = pd.to_datetime(out[c], errors="coerce")
            except Exception:
                pass

    return out


def render_download_bar(
    df_for_export: pd.DataFrame,
    *,
    key_prefix: str = "issues",
    caption: str = "",
    filename_prefix: str = "issues_filtradas",
    suffix: str = "",
) -> None:
    """Small reusable bar: download button + count/caption."""
    c1, c2 = st.columns([1, 3])
    with c1:
        download_button_for_df(
            df_for_export,
            key=f"{key_prefix}::download_csv",
            spec=CsvDownloadSpec(filename_prefix=filename_prefix),
            suffix=suffix,
            disabled=df_for_export is None or df_for_export.empty,
        )
    with c2:
        if caption:
            st.caption(caption)
        else:
            n = 0 if df_for_export is None else int(len(df_for_export))
            st.caption(f"{n} issues (según filtros actuales)")


def fig_to_html_bytes(fig: Any) -> bytes:
    if fig is None:
        return b""
    to_html = getattr(fig, "to_html", None)
    if not callable(to_html):
        return b""
    try:
        html_doc = to_html(include_plotlyjs="cdn", full_html=True)
    except Exception:
        return b""
    return str(html_doc).encode("utf-8", errors="replace")


def fig_to_svg_bytes(fig: Any) -> bytes:
    if fig is None:
        return b""
    to_image = getattr(fig, "to_image", None)
    if not callable(to_image):
        return b""
    try:
        svg = to_image(format="svg")
    except Exception:
        return b""
    if isinstance(svg, bytes):
        return svg
    return str(svg).encode("utf-8", errors="replace")


def figures_to_html_bytes(
    figures: Sequence[Any],
    *,
    title: str = "Export",
    subtitles: Optional[Sequence[str]] = None,
) -> bytes:
    """Build a single HTML document from multiple Plotly figures."""
    if not figures:
        return b""

    caps = list(subtitles or [])
    blocks: list[str] = []
    for i, fig in enumerate(figures):
        to_html = getattr(fig, "to_html", None)
        if not callable(to_html):
            continue
        try:
            block = to_html(include_plotlyjs=False, full_html=False)
        except Exception:
            continue
        cap = caps[i] if i < len(caps) else f"Chart {i + 1}"
        blocks.append(
            f"""
            <section class="panel">
              <h3>{html.escape(str(cap))}</h3>
              {block}
            </section>
            """
        )

    if not blocks:
        return b""

    doc = f"""<!doctype html>
<html lang="en">
  <head>
    <meta charset="utf-8" />
    <meta name="viewport" content="width=device-width, initial-scale=1" />
    <title>{html.escape(title)}</title>
    <script src="https://cdn.plot.ly/plotly-2.35.2.min.js"></script>
    <style>
      body {{
        margin: 0;
        padding: 22px;
        background: {BBVA_LIGHT.bg_light};
        color: {BBVA_LIGHT.ink};
        font-family: {BBVA_FONT_SANS};
      }}
      .wrap {{
        display: grid;
        grid-template-columns: 1fr;
        gap: 16px;
      }}
      .panel {{
        background: {BBVA_LIGHT.white};
        border: 1px solid rgba(17,25,45,0.14);
        border-radius: 16px;
        padding: 12px;
      }}
      h1 {{
        margin: 0 0 10px 0;
        font-size: 1.1rem;
        font-family: {BBVA_FONT_HEADLINE};
      }}
      h3 {{
        margin: 0 0 8px 0;
        font-size: 0.9rem;
        font-weight: 700;
        color: rgba(17,25,45,0.72);
        font-family: {BBVA_FONT_SANS};
      }}
    </style>
  </head>
  <body>
    <h1>{html.escape(title)}</h1>
    <div class="wrap">
      {''.join(blocks)}
    </div>
  </body>
</html>
"""
    return doc.encode("utf-8", errors="replace")


def _inject_minimal_export_css(scope_key: str) -> None:
    st.markdown(
        f"""
        <style>
          .st-key-{scope_key} {{
            margin-top: -0.10rem;
            margin-bottom: 0.12rem;
            padding-right: 0.34rem;
          }}
          .st-key-{scope_key} [data-testid="stHorizontalBlock"] {{
            justify-content: flex-end !important;
            align-items: center !important;
            gap: 0.32rem !important;
            margin: 0 !important;
            flex-wrap: wrap !important;
          }}
          .st-key-{scope_key} [data-testid="stColumn"] {{
            flex: 0 0 auto !important;
            width: auto !important;
            min-width: 0 !important;
            display: flex;
            justify-content: flex-end;
            align-items: center;
          }}
          .st-key-{scope_key} [data-testid="stColumn"] > div {{
            display: flex;
            justify-content: flex-end;
            align-items: center;
            width: auto !important;
          }}
          .st-key-{scope_key} .stDownloadButton {{
            width: auto !important;
            display: flex;
            justify-content: flex-end;
          }}
        </style>
        """,
        unsafe_allow_html=True,
    )


def render_minimal_export_actions(
    *,
    key_prefix: str,
    filename_prefix: str,
    suffix: str = "",
    csv_df: Optional[pd.DataFrame] = None,
    figure: Any = None,
    html_bytes: Optional[bytes] = None,
) -> None:
    """Minimal right-aligned exports (CSV + chart HTML when available)."""
    csv_safe = csv_df if isinstance(csv_df, pd.DataFrame) else pd.DataFrame()
    fig_html = html_bytes if html_bytes else fig_to_html_bytes(figure)

    has_csv = not csv_safe.empty
    has_html = bool(fig_html)
    if not has_csv and not has_html:
        return

    scope_key = f"{_safe_filename(key_prefix)}_mini_export"
    _inject_minimal_export_css(scope_key)

    buttons: list[dict[str, Any]] = []
    if has_csv:
        buttons.append(
            {
                "label": "Excel",
                "data": df_to_excel_bytes(csv_safe, sheet_name="Datos"),
                "file_name": _build_filename(filename_prefix, suffix=suffix, ext="xlsx"),
                "mime": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                "key": f"{key_prefix}::dl_csv_min",
                "disabled": False,
                "help": None,
            }
        )
    if has_html:
        buttons.append(
            {
                "label": "HTML",
                "data": fig_html,
                "file_name": _build_filename(filename_prefix, suffix=suffix, ext="html"),
                "mime": "text/html",
                "key": f"{key_prefix}::dl_html_min",
                "disabled": False,
                "help": None,
            }
        )

    with st.container(key=scope_key):
        cols = st.columns(len(buttons), gap="small")
        for col, btn in zip(cols, buttons):
            with col:
                st.download_button(
                    label=btn["label"],
                    data=btn["data"],
                    file_name=btn["file_name"],
                    mime=btn["mime"],
                    key=btn["key"],
                    width="content",
                    disabled=bool(btn.get("disabled", False)),
                    help=btn.get("help"),
                )
